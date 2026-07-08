"""Build dashboard.xlsx from audit.csv. Rerun anytime for fresh data:

    python scripts/build_dashboard.py

Pairs BUY/SELL audit rows into trades, then builds a 3-sheet workbook
(Summary KPIs + daily charts, Analysis breakdowns, Trades ledger).
All stats are Excel formulas over the Trades sheet.
"""
import csv
import datetime as dt
import os
import sys

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
AUDIT = os.path.join(ROOT, 'audit.csv')
OUT = os.path.join(ROOT, 'dashboard.xlsx')

# Rows logged before this date carry local CDT timestamps; add 1h to normalize to ET
ET_CUTOVER = dt.datetime(2026, 7, 5)

HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF')
HDR_FILL = PatternFill('solid', start_color='305496')
BODY_FONT = Font(name='Arial')
TITLE_FONT = Font(name='Arial', bold=True, size=14)
MONEY = '$#,##0.00;($#,##0.00);-'
PCT = '0.0%;(0.0%);-'


def exit_category(reason):
    r = reason.lower()
    if 'thesis invalidated' in r:
        return 'Invalidation'
    if 'hard stop' in r:
        return 'Hard Stop'
    if 'trailing stop after' in r:
        return 'Trail (new)'
    if 'dropped to 70%' in r or 'hard stop loss 50%' in r:
        return 'Old rule (removed)'
    return 'Other'


def orphan_row(ts, entry):
    return {
        'date': ts.date(), 'symbol': entry['Symbol'], 'direction': entry['Direction'],
        'entry_ts': ts, 'exit_ts': None,
        'entry_px': float(entry['Price']), 'exit_px': None, 'hold_min': None,
        'entry_hour': ts.hour,
        'adx': float(entry['ADX']) if entry.get('ADX', '').strip() else None,
        'adx_slope': None, 'pct': None, 'peak': None, 'pnl': None,
        'exit_cat': '', 'status': 'Orphaned/manual',
        'conviction': entry.get('Conviction', '') or '',
        'commission': None,
    }


def parse_trades():
    trades, open_pos = [], {}
    with open(AUDIT) as f:
        for row in csv.DictReader(f):
            ts = dt.datetime.strptime(row['Timestamp'], '%Y-%m-%d %H:%M:%S')
            if ts < ET_CUTOVER:
                ts += dt.timedelta(hours=1)  # CDT -> ET
            sym, action = row['Symbol'], row['Action']
            if action == 'BUY':
                if sym in open_pos:  # unclosed earlier BUY -> orphaned, don't lose it
                    old_ts, old_entry = open_pos[sym]
                    trades.append(orphan_row(old_ts, old_entry))
                open_pos[sym] = (ts, row)
            else:
                entry_ts, entry = open_pos.pop(sym, (None, {}))
                pnl = float(row['Dollar_PnL']) if row['Dollar_PnL'].strip() else None
                pct = float(row['Profit_Pct'].rstrip('%')) / 100 if row['Profit_Pct'].strip() else None
                peak = float(row['Peak_Pct'].rstrip('%')) / 100 if row.get('Peak_Pct', '').strip() else None
                slope = float(entry.get('ADX_Slope') or 0) if entry.get('ADX_Slope', '').strip() else None
                trades.append({
                    'date': ts.date(), 'symbol': sym, 'direction': row['Direction'],
                    'entry_ts': entry_ts, 'exit_ts': ts,
                    'entry_px': float(entry['Price']) if entry else None,
                    'exit_px': float(row['Price']),
                    'hold_min': round((ts - entry_ts).total_seconds() / 60) if entry_ts else None,
                    'entry_hour': entry_ts.hour if entry_ts else None,
                    'adx': float(entry['ADX']) if entry.get('ADX', '').strip() else None,
                    'adx_slope': slope, 'pct': pct, 'peak': peak, 'pnl': pnl,
                    'exit_cat': exit_category(row['Reason']), 'status': 'Closed',
                    'conviction': entry.get('Conviction', '') or '',
                    'commission': float(row['Commission']) if (row.get('Commission') or '').strip() else None,
                })
    for sym, (ts, entry) in open_pos.items():
        trades.append(orphan_row(ts, entry))
    trades.sort(key=lambda t: (t['date'], t['entry_ts'] or t['exit_ts']))
    return trades


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font, cell.fill = HDR_FONT, HDR_FILL
        cell.alignment = Alignment(horizontal='center')


def build():
    trades = parse_trades()
    wb = Workbook()

    # ── Trades sheet ────────────────────────────────────────────────────────
    tr = wb.create_sheet('Trades')
    cols = ['Date', 'Symbol', 'Direction', 'Entry (ET)', 'Exit (ET)', 'Hold (min)',
            'Entry Hr', 'Entry $', 'Exit $', 'ADX', 'ADX Slope', 'Peak %', 'P&L %',
            'P&L $', 'Exit Rule', 'Status', 'Conviction', 'Commission']
    tr.append(cols)
    style_header(tr, 1, len(cols))
    for t in trades:
        tr.append([
            t['date'], t['symbol'], t['direction'],
            t['entry_ts'].strftime('%m-%d %H:%M') if t['entry_ts'] else '',
            t['exit_ts'].strftime('%m-%d %H:%M') if t['exit_ts'] else '',
            t['hold_min'], t['entry_hour'], t['entry_px'], t['exit_px'],
            t['adx'], t['adx_slope'], t['peak'], t['pct'], t['pnl'],
            t['exit_cat'], t['status'], t['conviction'], t['commission'],
        ])
    n = len(trades) + 1
    for r in range(2, n + 1):
        tr.cell(row=r, column=12).number_format = PCT
        tr.cell(row=r, column=13).number_format = PCT
        tr.cell(row=r, column=14).number_format = MONEY
        tr.cell(row=r, column=18).number_format = MONEY
        for c in range(1, len(cols) + 1):
            tr.cell(row=r, column=c).font = BODY_FONT
    for i, w in enumerate([11, 8, 9, 12, 12, 9, 8, 8, 8, 7, 9, 8, 8, 11, 18, 16, 30, 11], 1):
        tr.column_dimensions[get_column_letter(i)].width = w

    # ── Daily aggregates (data block feeds Summary charts) ─────────────────
    dy = wb.create_sheet('Daily')
    dy.append(['Date', 'Net P&L $', 'Cumulative $', 'Trades', 'Wins', 'Losses'])
    style_header(dy, 1, 6)
    dates = sorted({t['date'] for t in trades if t['pnl'] is not None})
    for i, d in enumerate(dates, start=2):
        dy.cell(row=i, column=1, value=d.strftime('%m/%d'))
        ds = d.strftime('%Y-%m-%d')
        dy.cell(row=i, column=2, value=f'=SUMPRODUCT((TEXT(Trades!$A$2:$A${n},"yyyy-mm-dd")="{ds}")*Trades!$N$2:$N${n})')
        dy.cell(row=i, column=3, value=f'=SUM($B$2:B{i})')
        dy.cell(row=i, column=4, value=f'=SUMPRODUCT((TEXT(Trades!$A$2:$A${n},"yyyy-mm-dd")="{ds}")*(Trades!$P$2:$P${n}="Closed")*1)')
        dy.cell(row=i, column=5, value=f'=SUMPRODUCT((TEXT(Trades!$A$2:$A${n},"yyyy-mm-dd")="{ds}")*(Trades!$N$2:$N${n}>0)*1)')
        dy.cell(row=i, column=6, value=f'=D{i}-E{i}')
        dy.cell(row=i, column=2).number_format = MONEY
        dy.cell(row=i, column=3).number_format = MONEY
    for i, w in enumerate([10, 12, 13, 8, 7, 8], 1):
        dy.column_dimensions[get_column_letter(i)].width = w
    ndays = len(dates) + 1

    # ── Analysis sheet: symbol / exit rule / entry hour ─────────────────────
    an = wb.create_sheet('Analysis')
    an['A1'] = 'By Symbol'
    an['A1'].font = TITLE_FONT
    an.append(['Symbol', 'P&L $', 'Trades', 'Win rate'])
    style_header(an, 2, 4)
    for i, s in enumerate(sorted({t['symbol'] for t in trades}), start=3):
        an.cell(row=i, column=1, value=s)
        an.cell(row=i, column=2, value=f'=SUMIFS(Trades!$N$2:$N${n},Trades!$B$2:$B${n},A{i})')
        an.cell(row=i, column=3, value=f'=COUNTIFS(Trades!$B$2:$B${n},A{i},Trades!$P$2:$P${n},"Closed")')
        an.cell(row=i, column=4, value=f'=IF(C{i}=0,0,COUNTIFS(Trades!$B$2:$B${n},A{i},Trades!$N$2:$N${n},">0")/C{i})')
        an.cell(row=i, column=2).number_format = MONEY
        an.cell(row=i, column=4).number_format = PCT
    sym_end = 2 + len({t['symbol'] for t in trades})

    r0 = sym_end + 3
    an.cell(row=r0, column=1, value='By Exit Rule').font = TITLE_FONT
    cats = ['Invalidation', 'Hard Stop', 'Trail (new)', 'Old rule (removed)']
    hdr = r0 + 1
    for j, h in enumerate(['Exit Rule', 'P&L $', 'Count', 'Avg P&L $'], 1):
        an.cell(row=hdr, column=j, value=h)
    style_header(an, hdr, 4)
    for i, cat in enumerate(cats, start=hdr + 1):
        an.cell(row=i, column=1, value=cat)
        an.cell(row=i, column=2, value=f'=SUMIFS(Trades!$N$2:$N${n},Trades!$O$2:$O${n},A{i})')
        an.cell(row=i, column=3, value=f'=COUNTIF(Trades!$O$2:$O${n},A{i})')
        an.cell(row=i, column=4, value=f'=IF(C{i}=0,0,B{i}/C{i})')
        an.cell(row=i, column=2).number_format = MONEY
        an.cell(row=i, column=4).number_format = MONEY
    cat_hdr, cat_end = hdr, hdr + len(cats)

    r1 = cat_end + 3
    an.cell(row=r1, column=1, value='By Entry Hour (ET)').font = TITLE_FONT
    hhdr = r1 + 1
    for j, h in enumerate(['Hour', 'P&L $', 'Trades'], 1):
        an.cell(row=hhdr, column=j, value=h)
    style_header(an, hhdr, 3)
    hours = list(range(9, 16))
    for i, hh in enumerate(hours, start=hhdr + 1):
        an.cell(row=i, column=1, value=f'{hh}:00')
        an.cell(row=i, column=2, value=f'=SUMIFS(Trades!$N$2:$N${n},Trades!$G$2:$G${n},{hh})')
        an.cell(row=i, column=3, value=f'=COUNTIFS(Trades!$G$2:$G${n},{hh},Trades!$P$2:$P${n},"Closed")')
        an.cell(row=i, column=2).number_format = MONEY
    hr_hdr, hr_end = hhdr, hhdr + len(hours)

    r2 = hr_end + 3
    an.cell(row=r2, column=1, value='By Conviction Tier').font = TITLE_FONT
    chdr = r2 + 1
    for j, h in enumerate(['Tier', 'P&L $', 'Trades', 'Win rate'], 1):
        an.cell(row=chdr, column=j, value=h)
    style_header(an, chdr, 4)
    for i, tier in enumerate(['LOW', 'MEDIUM', 'HIGH'], start=chdr + 1):
        an.cell(row=i, column=1, value=tier)
        an.cell(row=i, column=2, value=f'=SUMIFS(Trades!$N$2:$N${n},Trades!$Q$2:$Q${n},A{i}&" *")')
        an.cell(row=i, column=3, value=f'=COUNTIFS(Trades!$Q$2:$Q${n},A{i}&" *",Trades!$P$2:$P${n},"Closed")')
        an.cell(row=i, column=4, value=f'=IF(C{i}=0,0,COUNTIFS(Trades!$Q$2:$Q${n},A{i}&" *",Trades!$N$2:$N${n},">0")/C{i})')
        an.cell(row=i, column=2).number_format = MONEY
        an.cell(row=i, column=4).number_format = PCT
    conv_hdr, conv_end = chdr, chdr + 3

    for i, w in enumerate([20, 12, 8, 10], 1):
        an.column_dimensions[get_column_letter(i)].width = w

    def bar(sheet, title, cat_ref, val_ref, anchor, width=13):
        ch = BarChart()
        ch.title, ch.height, ch.width, ch.style = title, 7.5, width, 10
        ch.add_data(val_ref, titles_from_data=True)
        ch.set_categories(cat_ref)
        ch.legend = None
        sheet.add_chart(ch, anchor)

    bar(an, 'P&L by Symbol', Reference(an, min_col=1, min_row=3, max_row=sym_end),
        Reference(an, min_col=2, min_row=2, max_row=sym_end), 'F2')
    bar(an, 'P&L by Exit Rule', Reference(an, min_col=1, min_row=cat_hdr + 1, max_row=cat_end),
        Reference(an, min_col=2, min_row=cat_hdr, max_row=cat_end), 'F18')
    bar(an, 'P&L by Entry Hour (ET)', Reference(an, min_col=1, min_row=hr_hdr + 1, max_row=hr_end),
        Reference(an, min_col=2, min_row=hr_hdr, max_row=hr_end), 'F34')
    bar(an, 'P&L by Conviction Tier', Reference(an, min_col=1, min_row=conv_hdr + 1, max_row=conv_end),
        Reference(an, min_col=2, min_row=conv_hdr, max_row=conv_end), 'F50')

    # ── Summary sheet ───────────────────────────────────────────────────────
    sm = wb.active
    sm.title = 'Summary'
    sm['A1'] = '0DTE Bot — Performance Dashboard'
    sm['A1'].font = Font(name='Arial', bold=True, size=16)
    sm['A2'] = f'Generated {dt.datetime.now():%Y-%m-%d %H:%M} — refresh: python scripts/build_dashboard.py'
    sm['A2'].font = Font(name='Arial', italic=True, size=9, color='808080')
    kpis = [
        ('Total P&L $', f'=SUM(Trades!$N$2:$N${n})', MONEY),
        ('Closed trades', f'=COUNTIF(Trades!$P$2:$P${n},"Closed")', '0'),
        ('Win rate', f'=COUNTIF(Trades!$N$2:$N${n},">0")/B5', PCT),
        ('Avg win $', f'=AVERAGEIF(Trades!$N$2:$N${n},">0")', MONEY),
        ('Avg loss $', f'=IF(COUNTIF(Trades!$N$2:$N${n},"<0")=0,0,AVERAGEIF(Trades!$N$2:$N${n},"<0"))', MONEY),
        ('Profit factor', f'=IF(SUMIF(Trades!$N$2:$N${n},"<0")=0,0,-SUMIF(Trades!$N$2:$N${n},">0")/SUMIF(Trades!$N$2:$N${n},"<0"))', '0.00'),
        ('Best day $', f'=MAX(Daily!$B$2:$B${ndays})', MONEY),
        ('Worst day $', f'=MIN(Daily!$B$2:$B${ndays})', MONEY),
        ('Commissions $', f'=SUM(Trades!$R$2:$R${n})', MONEY),
        ('Net after fees $', '=B4-B12', MONEY),
    ]
    for i, (label, formula, fmt) in enumerate(kpis, start=4):
        sm.cell(row=i, column=1, value=label).font = Font(name='Arial', bold=True)
        c = sm.cell(row=i, column=2, value=formula)
        c.number_format, c.font = fmt, BODY_FONT
    sm.column_dimensions['A'].width = 16
    sm.column_dimensions['B'].width = 14

    bar(sm, 'Daily Net P&L ($)', Reference(dy, min_col=1, min_row=2, max_row=ndays),
        Reference(dy, min_col=2, min_row=1, max_row=ndays), 'D2', width=15)
    eq = LineChart()
    eq.title, eq.height, eq.width, eq.style = 'Equity Curve (cumulative $)', 7.5, 15, 12
    eq.add_data(Reference(dy, min_col=3, min_row=1, max_row=ndays), titles_from_data=True)
    eq.set_categories(Reference(dy, min_col=1, min_row=2, max_row=ndays))
    eq.legend = None
    sm.add_chart(eq, 'D18')

    wb.save(OUT)
    print(f'Wrote {OUT}: {len(trades)} trades, {len(dates)} trading days')


if __name__ == '__main__':
    build()
